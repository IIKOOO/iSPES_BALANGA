from flask import Blueprint, send_file, session, make_response
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from io import BytesIO
from app import conn
from datetime import datetime, date
from collections import defaultdict
import os
import csv
import io

reports_bp = Blueprint('reports', __name__)    
    
@reports_bp.route('/download_final_spes_list_xlsx')
def download_final_spes_list_xlsx():
    # Query student data
    with conn.cursor() as cur:
        cur.execute("""
            SELECT
                s.student_id, s.last_name, s.first_name, s.middle_name, s.suffix, s.birth_date,
                s.sex, s.street_add, s.barangay, s.mobile_no, e.educational_attainment, s.iskolar_type
            FROM student_application s
            LEFT JOIN student_application_education e ON s.student_id = e.student_application_id
            WHERE s.is_pending = FALSE AND s.is_approved = TRUE
            ORDER BY s.last_name, s.first_name
        """)
        rows = cur.fetchall()

    # Calculate male, female, total
    male_count = sum(1 for r in rows if str(r[6]).strip().lower() == 'male')
    female_count = sum(1 for r in rows if str(r[6]).strip().lower() == 'female')
    total_count = len(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SPES Report"

    # Header: merge A1:I12
    ws.merge_cells('A1:I12')
    header_text = (
        "REPUBLIC OF THE PHILIPPINES\n"
        "DEPARTMENT OF LABOR AND EMPLOYMENT\n"
        "REGION OFFICE NO. III\n"
        "PUBLIC EMPLOYMENT SERVICE OFFICE\n"
        "CITY GOVERNMENT OF BALANGA\n"
        "SPECIAL PROGRAM FOR EMPLOYMENT OF STUDENTS (SPES)\n"
        "(RA 7323, as amended by RAs 9547 and 10917)"
    )
    ws['A1'] = header_text
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # Set font sizes for each line
    for i, (text, size) in enumerate([
        ("DEPARTMENT OF LABOR AND EMPLOYMENT", 16),
        ("REGION OFFICE NO. III", 12),
        ("PUBLIC EMPLOYMENT SERVICE OFFICE", 12),
        ("CITY GOVERNMENT OF BALANGA", 16),
        ("SPECIAL PROGRAM FOR EMPLOYMENT OF STUDENTS (SPES)", 16),
        ("(RA 7323, as amended by RAs 9547 and 10917)", 12)
    ]):
        ws['A1'].font = Font(size=16, bold=True)  # Set default font for merged cell

    # Insert logos
    logo_left = os.path.join('static', 'images', 'peso_balanga_logo.png')
    logo_right = os.path.join('static', 'images', 'spes_logo.png')
    if os.path.exists(logo_left):
        img_left = XLImage(logo_left)
        img_left.height = 100
        img_left.width = 100
        ws.add_image(img_left, 'A4')
    if os.path.exists(logo_right):
        img_right = XLImage(logo_right)
        img_right.height = 100
        img_right.width = 100
        ws.add_image(img_right, 'I4')
        
    ws.merge_cells('B13:H15')
    ws['B13'] = "FINAL SPES LIST REPORT"
    ws['B13'].font = Font(size=20, bold=True)
    ws['B13'].alignment = Alignment(horizontal='center', vertical='center')    

    # Summary rows: B13 to B15
    ws['G16'] = f"Total = {total_count}"
    ws['H16'] = f"Male = {male_count}"
    ws['I16'] = f"Female = {female_count}"
    for cell in ['G16', 'H16', 'I16']:
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')

    # Table header at row 17
    header_row = 17
    headers = [
        "No.", "SPES BENEFICIARY", "Student ID/SPES ID", "AGE", "SEX",
        "ADDRESS", "CONTACT NO.", "EDUCATIONAL LEVEL", "ISKOLAR TYPE"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="808080")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Borders
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Write student rows
    for idx, row in enumerate(rows, 1):
        # Full name: last_name, first_name (suffix), middle_name
        last_name = row[1] or ""
        first_name = row[2] or ""
        suffix = row[4] or ""
        middle_name = row[3] or ""
        full_name = f"{last_name}, {first_name}"
        if suffix:
            full_name += f" {suffix}"
        if middle_name:
            full_name += f" {middle_name}"

        # Age calculation
        birth_date = row[5]
        age = ""
        if isinstance(birth_date, (datetime, date)):
            today = date.today()
            age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))

        # Address
        street_add = row[7] or ""
        barangay = row[8] or ""
        address = f"{street_add}, {barangay}".strip(", ")

        # Educational Level
        educational_attainment = row[10] or ""

        # Iskolar Type
        iskolar_type = row[11] or ""

        # Table row
        table_row = [
            idx,
            full_name,
            row[0],  # student_id
            age,
            row[6],  # sex
            address,
            row[9],  # mobile_no
            educational_attainment,
            iskolar_type
        ]
        for col_num, value in enumerate(table_row, 1):
            cell = ws.cell(row=header_row + idx, column=col_num, value=value)
            # Borders for table
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            # Stretch columns for beneficiary and address
            if col_num in [2, 6]:
                ws.column_dimensions[get_column_letter(col_num)].width = 25
            else:
                ws.column_dimensions[get_column_letter(col_num)].width = 15

    # Save to BytesIO and send as response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        download_name="final_spes_list_report.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )    
    
@reports_bp.route('/download_student_payroll_xlsx')
def download_student_payroll_xlsx():
    with conn.cursor() as cur:
        cur.execute("""
            SELECT
                dtr_record_id, last_name, first_name, middle_name, suffix
            FROM student_dtr_records
            WHERE for_payroll = TRUE
            ORDER BY last_name, first_name
        """)
        rows = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Payroll List"

    # Header: merge A1:C10
    ws.merge_cells('A1:C10')
    header_text = (
        "REPUBLIC OF THE PHILIPPINES\n"
        "DEPARTMENT OF LABOR AND EMPLOYMENT\n"
        "REGION OFFICE NO. III\n"
        "PUBLIC EMPLOYMENT SERVICE OFFICE\n"
        "CITY GOVERNMENT OF BALANGA\n"
        "STUDENT PAYROLL LIST\n"
        "(RA 7323, as amended by RAs 9547 and 10917)"
    )
    ws['A1'] = header_text
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['A1'].font = Font(size=14, bold=True)

    # Insert logos (optional, adjust as needed)
    logo_left = os.path.join('static', 'images', 'peso_balanga_logo.png')
    logo_right = os.path.join('static', 'images', 'spes_logo.png')
    if os.path.exists(logo_left):
        img_left = XLImage(logo_left)
        img_left.height = 80
        img_left.width = 80
        ws.add_image(img_left, 'A4')
    if os.path.exists(logo_right):
        img_right = XLImage(logo_right)
        img_right.height = 80
        img_right.width = 80
        # Place at the end center of the header (C7 for center of A1:C10)
        ws.add_image(img_right, 'C4')

    # Title: merge A11:C11, text "STUDENT PAYROLL LIST"
    ws.merge_cells('A11:C11')
    ws['A11'] = "STUDENT PAYROLL LIST"
    ws['A11'].font = Font(size=18, bold=True)
    ws['A11'].alignment = Alignment(horizontal='center', vertical='center')

    # Table header at row 13 (was 16)
    header_row = 13
    headers = ["No.", "Name", "Cashier Number"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="808080")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # Set column widths
        if col_num == 1:
            ws.column_dimensions[get_column_letter(col_num)].width = 10  # No.
        else:
            ws.column_dimensions[get_column_letter(col_num)].width = 50  # Name, Cashier Number (longer)

    # Write student rows
    for idx, row in enumerate(rows, 1):
        last_name = row[1] or ""
        first_name = row[2] or ""
        middle_name = row[3] or ""
        suffix = row[4] or ""
        full_name = f"{last_name}, {first_name}"
        if suffix:
            full_name += f" {suffix}"
        if middle_name:
            full_name += f" {middle_name}"

        ws.cell(row=header_row + idx, column=1, value=idx).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=header_row + idx, column=2, value=full_name)
        cashier_cell = ws.cell(row=header_row + idx, column=3, value="")
        cashier_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Borders for all cells
        for col_num in range(1, 4):
            cell = ws.cell(row=header_row + idx, column=col_num)
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Save to BytesIO and send as response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        download_name="student_payroll_list.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@reports_bp.route('/download_student_registration_csv')
def download_student_registration_csv():
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    s.student_id, s.student_category, s.iskolar_type, s.status_of_student, s.last_name, s.first_name, s.middle_name, s.suffix,
                    s.birth_date, s.birth_place, s.citizenship, s.mobile_no, s.email, s.civil_status, s.sex, s.disabilities, s.socmed,
                    s.participation_count, s.belongs_in_group, s.barangay, s.street_add, s.name_of_beneficiary, s.relationship_to_beneficiary,
                    s.parents_status, s.time_stamp,
                    e.school_name, e.educational_attainment, e.senior_high_strand, e.college_course, e.attendance_date,
                    p.living_with, p.guardian_full_name, p.guardian_contact_no, p.guardian_birth_date, p.guardian_occupation,
                    p.relationship_with_guardian, p.guardian_tin_no, p.father_full_name, p.father_contact_no, p.father_birth_date,
                    p.father_occupation, p.father_tin_no, p.mother_full_name, p.mother_contact_no, p.mother_birth_date,
                    p.mother_occupation, p.mother_tin_no
                FROM student_application s
                LEFT JOIN student_application_education e ON s.student_id = e.student_application_id
                LEFT JOIN student_application_parents_guardians p ON s.student_id = p.student_application_id
                WHERE s.is_pending = FALSE AND s.is_approved = FALSE
            """)
            rows = cur.fetchall()
        header = [
            "Student ID", "Student Category", "Iskolar Type", "Status of Student", "Last Name", "First Name", "Middle Name", "Suffix",
            "Birth Date", "Birth Place", "Citizenship", "Mobile No", "Email", "Civil Status", "Sex", "Disabilities", "Social Media",
            "Participation Count", "Belongs in Group", "Barangay", "Street Address", "Name of Beneficiary", "Relationship to Beneficiary",
            "Parents Status", "Timestamp",
            "School Name", "Educational Attainment", "Senior High Strand", "College Course", "Attendance Date",
            "Living With", "Guardian Full Name", "Guardian Contact No", "Guardian Birth Date", "Guardian Occupation",
            "Relationship With Guardian", "Guardian TIN No", "Father Full Name", "Father Contact No", "Father Birth Date",
            "Father Occupation", "Father TIN No", "Mother Full Name", "Mother Contact No", "Mother Birth Date",
            "Mother Occupation", "Mother TIN No"
        ]
        si = io.StringIO()
        writer = csv.writer(si)
        writer.writerow(header)
        import datetime
        for row in rows:
            csv_row = list(row)
            # Format dates safely
            def format_date(val, fmt='%Y-%m-%d'):
                if isinstance(val, (datetime.date, datetime.datetime)):
                    return val.strftime(fmt)
                return val if val else ''
            csv_row[8] = format_date(csv_row[8])  # Birth Date
            csv_row[23] = format_date(csv_row[23], '%Y-%m-%d %H:%M:%S')  # Timestamp
            csv_row[32] = format_date(csv_row[32])  # Attendance Date
            csv_row[34] = format_date(csv_row[34])  # Guardian Birth Date
            csv_row[37] = format_date(csv_row[37])  # Father Birth Date
            csv_row[42] = format_date(csv_row[42])  # Mother Birth Date
            writer.writerow(csv_row)
        output = si.getvalue()
        si.close()

        return make_response(
            output,
            200,
            {
                'Content-Type': 'text/csv',
                'Content-Disposition': 'attachment; filename=student_registration.csv'
            }
        )
    except Exception as e:
        return str(e), 500
    
@reports_bp.route('/download_pending_student_registration_csv')
def download_pending_student_registration_csv():
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    s.student_id, s.student_category, s.iskolar_type, s.status_of_student, s.last_name, s.first_name, s.middle_name, s.suffix,
                    s.birth_date, s.birth_place, s.citizenship, s.mobile_no, s.email, s.civil_status, s.sex, s.disabilities, s.socmed,
                    s.participation_count, s.belongs_in_group, s.barangay, s.street_add, s.name_of_beneficiary, s.relationship_to_beneficiary,
                    s.parents_status, s.time_stamp,
                    e.school_name, e.educational_attainment, e.senior_high_strand, e.college_course, e.attendance_date,
                    p.living_with, p.guardian_full_name, p.guardian_contact_no, p.guardian_birth_date, p.guardian_occupation,
                    p.relationship_with_guardian, p.guardian_tin_no, p.father_full_name, p.father_contact_no, p.father_birth_date,
                    p.father_occupation, p.father_tin_no, p.mother_full_name, p.mother_contact_no, p.mother_birth_date,
                    p.mother_occupation, p.mother_tin_no
                FROM student_application s
                LEFT JOIN student_application_education e ON s.student_id = e.student_application_id
                LEFT JOIN student_application_parents_guardians p ON s.student_id = p.student_application_id
                WHERE s.is_pending = TRUE AND s.is_approved = FALSE
            """)
            rows = cur.fetchall()
        header = [
            "Student ID", "Student Category", "Iskolar Type", "Status of Student", "Last Name", "First Name", "Middle Name", "Suffix",
            "Birth Date", "Birth Place", "Citizenship", "Mobile No", "Email", "Civil Status", "Sex", "Disabilities", "Social Media",
            "Participation Count", "Belongs in Group", "Barangay", "Street Address", "Name of Beneficiary", "Relationship to Beneficiary",
            "Parents Status", "Timestamp",
            "School Name", "Educational Attainment", "Senior High Strand", "College Course", "Attendance Date",
            "Living With", "Guardian Full Name", "Guardian Contact No", "Guardian Birth Date", "Guardian Occupation",
            "Relationship With Guardian", "Guardian TIN No", "Father Full Name", "Father Contact No", "Father Birth Date",
            "Father Occupation", "Father TIN No", "Mother Full Name", "Mother Contact No", "Mother Birth Date",
            "Mother Occupation", "Mother TIN No"
        ]
        si = io.StringIO()
        writer = csv.writer(si)
        writer.writerow(header)
        import datetime
        for row in rows:
            csv_row = list(row)
            # Format dates safely
            def format_date(val, fmt='%Y-%m-%d'):
                if isinstance(val, (datetime.date, datetime.datetime)):
                    return val.strftime(fmt)
                return val if val else ''
            csv_row[8] = format_date(csv_row[8])  # Birth Date
            csv_row[23] = format_date(csv_row[23], '%Y-%m-%d %H:%M:%S')  # Timestamp
            csv_row[32] = format_date(csv_row[32])  # Attendance Date
            csv_row[34] = format_date(csv_row[34])  # Guardian Birth Date
            csv_row[37] = format_date(csv_row[37])  # Father Birth Date
            csv_row[42] = format_date(csv_row[42])  # Mother Birth Date
            writer.writerow(csv_row)
        output = si.getvalue()
        si.close()

        return make_response(
            output,
            200,
            {
                'Content-Type': 'text/csv',
                'Content-Disposition': 'attachment; filename=student_registration.csv'
            }
        )
    except Exception as e:
        return str(e), 500

@reports_bp.route('/download_final_spes_list_csv')
def download_final_spes_list_csv():
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    s.student_id, s.student_category, s.iskolar_type, s.status_of_student, s.last_name, s.first_name, s.middle_name, s.suffix,
                    s.birth_date, s.birth_place, s.citizenship, s.mobile_no, s.email, s.civil_status, s.sex, s.disabilities, s.socmed,
                    s.participation_count, s.belongs_in_group, s.barangay, s.street_add, s.name_of_beneficiary, s.relationship_to_beneficiary,
                    s.parents_status, s.time_stamp,
                    e.school_name, e.educational_attainment, e.senior_high_strand, e.college_course, e.attendance_date,
                    p.living_with, p.guardian_full_name, p.guardian_contact_no, p.guardian_birth_date, p.guardian_occupation,
                    p.relationship_with_guardian, p.guardian_tin_no, p.father_full_name, p.father_contact_no, p.father_birth_date,
                    p.father_occupation, p.father_tin_no, p.mother_full_name, p.mother_contact_no, p.mother_birth_date,
                    p.mother_occupation, p.mother_tin_no
                FROM student_application s
                LEFT JOIN student_application_education e ON s.student_id = e.student_application_id
                LEFT JOIN student_application_parents_guardians p ON s.student_id = p.student_application_id
                WHERE s.is_pending = FALSE AND s.is_approved = TRUE
            """)
            rows = cur.fetchall()
        header = [
            "Student ID", "Student Category", "Iskolar Type", "Status of Student", "Last Name", "First Name", "Middle Name", "Suffix",
            "Birth Date", "Birth Place", "Citizenship", "Mobile No", "Email", "Civil Status", "Sex", "Disabilities", "Social Media",
            "Participation Count", "Belongs in Group", "Barangay", "Street Address", "Name of Beneficiary", "Relationship to Beneficiary",
            "Parents Status", "Timestamp",
            "School Name", "Educational Attainment", "Senior High Strand", "College Course", "Attendance Date",
            "Living With", "Guardian Full Name", "Guardian Contact No", "Guardian Birth Date", "Guardian Occupation",
            "Relationship With Guardian", "Guardian TIN No", "Father Full Name", "Father Contact No", "Father Birth Date",
            "Father Occupation", "Father TIN No", "Mother Full Name", "Mother Contact No", "Mother Birth Date",
            "Mother Occupation", "Mother TIN No"
        ]
        si = io.StringIO()
        writer = csv.writer(si)
        writer.writerow(header)
        import datetime
        for row in rows:
            csv_row = list(row)
            # Format dates safely
            def format_date(val, fmt='%Y-%m-%d'):
                if isinstance(val, (datetime.date, datetime.datetime)):
                    return val.strftime(fmt)
                return val if val else ''
            csv_row[8] = format_date(csv_row[8])  # Birth Date
            csv_row[23] = format_date(csv_row[23], '%Y-%m-%d %H:%M:%S')  # Timestamp
            csv_row[32] = format_date(csv_row[32])  # Attendance Date
            csv_row[34] = format_date(csv_row[34])  # Guardian Birth Date
            csv_row[37] = format_date(csv_row[37])  # Father Birth Date
            csv_row[42] = format_date(csv_row[42])  # Mother Birth Date
            writer.writerow(csv_row)
        output = si.getvalue()
        si.close()

        return make_response(
            output,
            200,
            {
                'Content-Type': 'text/csv',
                'Content-Disposition': 'attachment; filename=student_registration.csv'
            }
        )
    except Exception as e:
        return str(e), 500    

@reports_bp.route('/download_student_dtr_records_csv')
def download_student_dtr_records_csv():
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT dtr_record_id, last_name, first_name, middle_name, suffix, email, student_category, mobile_no, birth_date, total_worked_hours, starting_date, end_date
                FROM student_dtr_records
                WHERE for_payroll = FALSE
            """)
            rows = cur.fetchall()
        header = [
            "Student ID", "Last Name", "First Name", "Middle Name", "Suffix", "Email", "Student Category", "Mobile No", "Birth Date",
            "Total Lates", "Total On-time", "Total Worked Hours", "Start Date", "End Date"
        ]
        si = io.StringIO()
        writer = csv.writer(si)
        writer.writerow(header)
        import datetime
        for row in rows:
            csv_row = list(row)
            def format_date(val, fmt='%Y-%m-%d'):
                if isinstance(val, (datetime.date, datetime.datetime)):
                    return val.strftime(fmt)
                return val if val else ''
            # Safely format only if index exists
            for idx, fmt in [(8, '%Y-%m-%d'), (12, '%Y-%m-%d'), (13, '%Y-%m-%d')]:
                if len(csv_row) > idx:
                    csv_row[idx] = format_date(csv_row[idx], fmt)
            writer.writerow(csv_row)
        output = si.getvalue()
        si.close()

        return make_response(
            output,
            200,
            {
                'Content-Type': 'text/csv',
                'Content-Disposition': 'attachment; filename=student_dtr_records.csv'
            }
        )
    except Exception as e:
        return str(e), 500

@reports_bp.route('/download_student_payroll_csv')
def download_student_payroll_csv():
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT dtr_record_id, last_name, first_name, middle_name, suffix, email, student_category, mobile_no, birth_date, total_worked_hours, starting_date, end_date, is_paid
                FROM student_dtr_records
                WHERE for_payroll = TRUE
            """)
            rows = cur.fetchall()
        header = [
            "Student ID", "Last Name", "First Name", "Middle Name", "Suffix", "Email", "Student Category", "Mobile No", "Birth Date", "Total Worked Hours", "Start Date", "End Date", "Status"
        ]
        si = io.StringIO()
        writer = csv.writer(si)
        writer.writerow(header)
        import datetime
        for row in rows:
            csv_row = list(row)
            def format_date(val, fmt='%Y-%m-%d'):
                if isinstance(val, (datetime.date, datetime.datetime)):
                    return val.strftime(fmt)
                return val if val else ''
            # Safely format only if index exists
            for idx, fmt in [(8, '%Y-%m-%d'), (12, '%Y-%m-%d'), (13, '%Y-%m-%d')]:
                if len(csv_row) > idx:
                    csv_row[idx] = format_date(csv_row[idx], fmt)
            writer.writerow(csv_row)
        output = si.getvalue()
        si.close()

        return make_response(
            output,
            200,
            {
                'Content-Type': 'text/csv',
                'Content-Disposition': 'attachment; filename=student_payroll.csv'
            }
        )
    except Exception as e:
        return str(e), 500
    
@reports_bp.route('/download_csc_dtr_xlsx/<int:dtr_record_id>')
def download_csc_dtr_xlsx(dtr_record_id):
    # --- Fetch student and DTR data ---
    with conn.cursor() as cur:
        # Get student info
        cur.execute("""
            SELECT last_name, first_name, middle_name, suffix, starting_date, end_date
            FROM student_dtr_records WHERE dtr_record_id = %s
        """, (dtr_record_id,))
        student = cur.fetchone()
        if not student:
            return "Student not found", 404
        last_name, first_name, middle_name, suffix, starting_date, end_date = student

        # Get all DTR entries for this student
        cur.execute("""
            SELECT date, time_in_am, time_out_am, time_in_pm, time_out_pm
            FROM student_dtr WHERE student_id = %s
            ORDER BY date
        """, (dtr_record_id,))
        dtr_rows = cur.fetchall()

    # --- Group DTRs by (year, month) ---
    dtr_by_month = defaultdict(dict)
    for row in dtr_rows:
        date, am_in, am_out, pm_in, pm_out = row
        month_key = (date.year, date.month)
        dtr_by_month[month_key][date.day] = {
            "am_in": am_in,
            "am_out": am_out,
            "pm_in": pm_in,
            "pm_out": pm_out,
            "under_hours": 0,
            "under_minutes": 0
        }

    #helper to format time ---
    def format_time(dt, military=False):
        if not dt:
            return ''
        return dt.strftime("%H:%M:%S") if military else dt.strftime("%I:%M %p")

    #Prepare workbook ---
    wb = Workbook()
    first_sheet = True

    for (year, month), dtr_by_day in sorted(dtr_by_month.items()):
        # Create new sheet (or reuse first one)
        if first_sheet:
            ws = wb.active
            first_sheet = False
        else:
            ws = wb.create_sheet()
        ws.title = f"{datetime(year, month, 1).strftime('%B %Y')}"

        # --- 4. Header and meta ---
        ws.merge_cells('A1:H1')
        ws['A1'] = "Civil Service Form No. 48"
        ws['A1'].font = Font(italic=True, size=10)
        ws['A1'].alignment = Alignment(horizontal='left')

        ws.merge_cells('A2:H2')
        ws['A2'] = "DAILY TIME RECORD"
        ws['A2'].font = Font(bold=True, size=16)
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A3:H3')
        full_name = f"{last_name}, {first_name} {middle_name or ''} {suffix or ''}".strip()
        ws['A3'] = f"(Name): {full_name}"
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A3'].font = Font(italic=True)

        ws['A4'] = f"For the month of {datetime(year, month, 1).strftime('%B')}, {year}"
        ws['E4'] = "Regular Days ______"
        ws['G4'] = "Saturdays ______"

        # --- 5. Table headers ---
        ws['A6'] = "Days"
        ws.merge_cells('B6:C6')
        ws['B6'] = "A.M."
        ws.merge_cells('D6:E6')
        ws['D6'] = "P.M."
        ws.merge_cells('F6:G6')
        ws['F6'] = "UNDER TIME"
        ws['H6'] = ""
        ws['A7'] = ""
        ws['B7'] = "ARRIVAL"
        ws['C7'] = "DEPARTURE"
        ws['D7'] = "ARRIVAL"
        ws['E7'] = "DEPARTURE"
        ws['F7'] = "Hours"
        ws['G7'] = "Minutes"
        ws['H7'] = ""

        #Fill DTR rows (1-31) ---
        thin = Side(border_style="thin", color="000000")
        total_hours = 0
        total_minutes = 0

        for i in range(1, 32):
            row_idx = 7 + i
            ws[f'A{row_idx}'] = i
            dtr = dtr_by_day.get(i, {})
            ws[f'B{row_idx}'] = format_time(dtr.get('am_in'))
            ws[f'C{row_idx}'] = format_time(dtr.get('am_out'))
            ws[f'D{row_idx}'] = format_time(dtr.get('pm_in'))
            ws[f'E{row_idx}'] = format_time(dtr.get('pm_out'))
            ws[f'F{row_idx}'] = dtr.get('under_hours') or ''
            ws[f'G{row_idx}'] = dtr.get('under_minutes') or ''

            # Sum total under time
            try:
                total_hours += int(dtr.get('under_hours') or 0)
                total_minutes += int(dtr.get('under_minutes') or 0)
            except Exception:
                pass

            # Borders
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).border = Border(
                    top=thin, left=thin, right=thin, bottom=thin
                )
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center', vertical='center')

        #Total row ---
        ws[f'A{39}'] = "TOTAL"
        ws[f'F{39}'] = total_hours + (total_minutes // 60)
        ws[f'G{39}'] = total_minutes % 60
        for col in range(1, 8):
            ws.cell(row=39, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            ws.cell(row=39, column=col).alignment = Alignment(horizontal='center', vertical='center')

        #Certification ---
        ws.merge_cells('A41:G41')
        ws['A41'] = (
            "I CERTIFY on my honor that the above is a true and correct report of the hours of work performed, "
        )
        ws.merge_cells('A42:G42')
        ws['A42'] = (
            "record of which was made daily at the time of arrival and departure from office."
        )
        ws['A41'].alignment = Alignment(wrap_text=True)
        ws['A42'].alignment = Alignment(wrap_text=True)

        #Adjust column widths ---
        ws.column_dimensions['A'].width = 5
        for col in 'BCDEFG':
            ws.column_dimensions[col].width = 14

    #Save and send file ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"CSC_Form_48_DTR_{full_name.replace(' ', '_')}.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@reports_bp.route('/download_gsis_report_xlsx')
def download_gsis_report_xlsx():
    with conn.cursor() as cur:
        cur.execute("""
            SELECT
                s.student_id, s.last_name, s.first_name, s.middle_name, s.suffix, s.email, s.student_category, s.mobile_no, s.birth_date,
                s.iskolar_type, s.sex, e.educational_attainment, s.street_add, s.barangay, s.name_of_beneficiary
            FROM student_application s
            LEFT JOIN student_application_education e ON s.student_id = e.student_application_id
            WHERE s.is_approved = TRUE
            ORDER BY s.last_name, s.first_name
        """)
        rows = cur.fetchall()

    # Count male/female/total
    male_count = sum(1 for r in rows if str(r[10]).strip().lower() == 'male')
    female_count = sum(1 for r in rows if str(r[10]).strip().lower() == 'female')
    total_count = len(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GSIS Placement Report"

    # Header: merge A1:P12
    ws.merge_cells('A1:P12')
    header_text = (
        "REPUBLIC OF THE PHILIPPINES\n"
        "DEPARTMENT OF LABOR AND EMPLOYMENT\n"
        "REGION OFFICE NO. III\n"
        "PUBLIC EMPLOYMENT SERVICE OFFICE\n"
        "CITY GOVERNMENT OF BALANGA\n"
        "PLACEMENT REPORT CUM GSIS INSURANCE COVERAGE\n"
        "(RA 7323, as amended by RAs 9547 and 10917)"
    )
    ws['A1'] = header_text
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['A1'].font = Font(size=16, bold=True, name='Calibri')

    # Insert logos
    logo_left = os.path.join('static', 'images', 'peso_balanga_logo.png')
    logo_right = os.path.join('static', 'images', 'spes_logo.png')
    if os.path.exists(logo_left):
        img_left = XLImage(logo_left)
        img_left.height = 100
        img_left.width = 100
        ws.add_image(img_left, 'A4')
    if os.path.exists(logo_right):
        img_right = XLImage(logo_right)
        img_right.height = 100
        img_right.width = 100
        ws.add_image(img_right, 'P4')

    # Title: merge A13:P13, text "PLACEMENT REPORT CUM GSIS INSURANCE COVERAGE"
    ws.merge_cells('A13:P13')
    ws['A13'] = "PLACEMENT REPORT CUM GSIS INSURANCE COVERAGE"
    ws['A13'].font = Font(size=18, bold=True, name='Calibri')
    ws['A13'].alignment = Alignment(horizontal='center', vertical='center')

    # Employer Info: merge B14-G16
    ws.merge_cells('B14:G16')
    ws['B14'] = (
        "Name of Establishment/Employer: LGU BALANGA CITY\n"
        "Address: 2ND FLOOR NEGOSYO CENTER, ST. JOSEPH ST., POBLACION, BALANGA CITY, BATAAN\n"
        "Business Activity: GOVERNMENT"
    )
    ws['B14'].font = Font(size=11, bold=True, name='Calibri')
    ws['B14'].alignment = Alignment(wrap_text=True, vertical='top')

    # Industry Code: merge H16-J16
    ws.merge_cells('H16:J16')
    ws['H16'] = "Industry Code:_________"
    ws['H16'].font = Font(size=11, bold=True, name='Calibri')
    ws['H16'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='top')

    # Beneficiaries/Contact: merge K14:P16
    ws.merge_cells('K14:P16')
    ws['K14'] = (
        f"Number of Beneficiaries: Female: {female_count} + Male: {male_count} = {total_count}\n"
        "Contact Person: NORIEL D. DACION JR/ City Government Department Head I (PESO Manager)\n"
        "Tel. No./Mobile No.: 237-0718"
    )
    ws['K14'].font = Font(size=11, bold=True, name='Calibri')
    ws['K14'].alignment = Alignment(wrap_text=True, vertical='top')

    # Table header at row 18
    header_row = 18
    headers = [
        "No.",
        "SPES BENEFICIARY",
        "Student ID",
        "AGE",
        "SEX",
        "ADDRESS",
        "Contact No.",
        "Student/ OSY/\nDependent of\nDisplaced Worker",
        "EDUCATIONAL LEVEL",
        "New/ SPES Baby",
        "Occupational Code\n& Position",
        "Wage Rate\nper Day",
        "EMPLOYMENT PERIOD",
        "Total Amount to be\nearned/received\nas salary/wages",
        "GSIS Policy\nNo.",
        "GSIS BENEFICIARY"
    ]
    ws.row_dimensions[header_row].height = 50  # Make header row taller

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        # Set font size 9 for columns 8, 11, 12, 14, 15
        if col_num in [8, 11, 12, 14, 15]:
            cell.font = Font(bold=True, color="FFFFFF", name='Calibri', size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            cell.font = Font(bold=True, color="FFFFFF", name='Calibri')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="808080")
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # Stretch columns for name, address, educational level, beneficiary
        if col_num == [1, 15]:
            ws.column_dimensions[get_column_letter(col_num)].width = 5
        if col_num in [2, 10, 11, 13, 16]:
            ws.column_dimensions[get_column_letter(col_num)].width = 25
        if col_num in [3, 4, 5]:
            ws.column_dimensions[get_column_letter(col_num)].width = 10
        if col_num in [6, 9]:
            ws.column_dimensions[get_column_letter(col_num)].width = 45
        if col_num in [7, 8, 14]:
            ws.column_dimensions[get_column_letter(col_num)].width = 15

    # Write student rows
    for idx, row in enumerate(rows, 1):
        # Full name: last_name, first_name (suffix), middle_name
        last_name = row[1] or ""
        first_name = row[2] or ""
        middle_name = row[3] or ""
        suffix = row[4] or ""
        full_name = f"{last_name}, {first_name}"
        if suffix:
            full_name += f" {suffix}"
        if middle_name:
            full_name += f" {middle_name}"

        # Age calculation
        birth_date = row[8]
        age = ""
        if isinstance(birth_date, (datetime, date)):
            today = date.today()
            age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))

        # Address: street_add, barangay
        street_add = row[12] or ""
        barangay = row[13] or ""
        address = f"{street_add}, {barangay}".strip(", ")

        # Contact No.
        contact_no = row[7] or ""

        # Educational Level (educational_attainment)
        educational_attainment = row[11] or ""

        # Iskolar Type
        iskolar_type = row[9] or ""

        # Sex
        sex = row[10] or ""

        # GSIS Beneficiary
        gsis_beneficiary = row[14] or ""

        # Table row
        table_row = [
            idx,
            full_name,
            row[0],  # student_id
            age,
            sex,
            address,
            contact_no,
            "",  # Student/ OSY/ Dependent of Displaced Worker (blank)
            educational_attainment,
            iskolar_type,  # New/ SPES Baby
            "",  # Occupational Code & Position (blank)
            "",  # Wage Rate per Day (blank)
            "",  # EMPLOYMENT PERIOD (blank)
            "",  # Total Amount to be earned/received as salary/wages (blank)
            "",  # GSIS Policy No. (blank)
            gsis_beneficiary  # GSIS BENEFICIARY (from name_of_beneficiary)
        ]
        for col_num, value in enumerate(table_row, 1):
            cell = ws.cell(row=header_row + idx, column=col_num, value=value)
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.font = Font(name='Calibri')
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Append contents after the table ---
    last_table_row = header_row + len(rows)
    note_row = last_table_row + 2

    # Merge A:P for the note row so it spans the whole table width
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=16)  # A:P

    note_cell = ws.cell(row=note_row, column=1)
    note_cell.value = "Note: This form shall be accomplished by the Public Employment Service Office to be submitted to the DOLE Regional Office at least ten (10) days prior to the date of employment."
    note_cell.font = Font(italic=True, size=8)
    note_cell.alignment = Alignment(wrap_text=False, horizontal='left', vertical='top')

    # Two rows after note, merge A-E and L-P for 7 rows each
    prepared_start = note_row + 2
    prepared_end = prepared_start + 6
    submitted_start = prepared_start
    submitted_end = prepared_end

    ws.merge_cells(start_row=prepared_start, start_column=1, end_row=prepared_end, end_column=5)  # A-E
    ws.merge_cells(start_row=submitted_start, start_column=12, end_row=submitted_end, end_column=16)  # L-P

    # Prepared By block
    prepared_text = (
        "Prepared By:\n\n"
        "_____________________________________________________________\n"
        "                                        Signature over Printed Name\n\n"
        "_____________________________________________________________\n"
        "                                                          Date"
    )
    cell = ws.cell(row=prepared_start, column=1, value=prepared_text)
    cell.font = Font(size=11)
    cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Submitted By block
    submitted_text = (
        "Submitted By:\n\n"
        "_____________________________________________________________\n"
        "                                        Signature over Printed Name\n\n"
        "_____________________________________________________________\n"
        "                                                          Date"
    )
    cell = ws.cell(row=submitted_start, column=12, value=submitted_text)
    cell.font = Font(size=11)
    cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Save to BytesIO and send as response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        download_name="gsis_placement_report.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )